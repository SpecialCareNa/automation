#coding=utf8
import openpyxl
import sys
import os

base_path = os.getcwd()
sys.path.append(base_path)

# open_excel = openpyxl.load_workbook(base_path+"\\case\\rqy.exml")
# sheet_name = open_excel.sheetnames
# excel_value = open_excel[sheet_name[0]]



class HandExcel:

    #加载excel
    def load_excel(self):
        path = 'C:/Users/Administrator/PycharmProjects/requests/case/rqy.xlsx'
        open_excel = openpyxl.load_workbook(path)
        # open_excel = openpyxl.load_workbook(base_path + '/rqy.xlsx')
        return open_excel

    #加载所有sheet的内容
    def get_sheet_data(self,index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    #获取单元格内容
    def get_cell_value(self,row,cols):
        data = self.get_sheet_data().cell(row=row,column=cols).value
        print(data)
        return data

    #获取行数
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row

    #获取某行内容
    def get_rows_value(self,row):
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

excel_data =  HandExcel()

if __name__ == '__main__':
    handle = HandExcel()
    print(handle.get_cell_value(1, 1))
    print(handle.get_rows_value(1))
