# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\Administrator\\PycharmProjects\\requests\\case\\rqy.xlsx')

ws = wb.active
cols = []
cols = []
for col in ws.iter_cols():
    cols.append(col)

print (cols)  # 所有列
print (cols[0])  # 获取第一列
print (cols[0][0]) # 获取第一列的第一行的单元格对象
print (cols[0][0].value)  # 获取第一列的第一行的值

print ("*" * 30)
print (cols[len(cols) - 1])  # 获取最后一列
print (cols[len(cols) - 1][len(cols[0]) - 1])  # 获取最后一列的最后一行的单元格对象
print (cols[len(cols) - 1][len(cols[0]) - 1].value)  # 获取最后一列的最后一行的单元格对象的值